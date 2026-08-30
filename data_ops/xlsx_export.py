"""XLSX export — DISPLAY ONLY.

Policy (docs/DATA_CENTER.md §1): Excel is a human view, never a source of
truth.  Values are written as displayed (numbers/dates as native cells, text
as text) and the workbook opens with ``__AMS_META__`` so a recipient can tell
which app/version produced it.  This module never *reads* workbooks back into
the database — JSON is the only restore/merge transport.
"""
from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path

from data_ops.constants import ARCHIVE_KIND, FORMAT_VERSION
from data_ops.portable import table_columns, tables_of

SHEET_MAX = 31
META_NAME = "__AMS_META__"


def _sheet_name(raw: str) -> str:
    """Excel sheet names: <=31 chars, no []:*?/\\ ."""
    cleaned = "".join("_" if ch in "[]:*?/\\" else ch for ch in str(raw))
    return (cleaned[:SHEET_MAX] or "Sheet").strip()


def export_xlsx(conn: sqlite3.Connection, dest: str | Path, *, tables: list[str] | None = None) -> dict:
    """Write a multi-sheet workbook; return {path, tables, rows}."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    names = tables_of(conn) if tables is None else [t for t in tables if t in set(tables_of(conn))]

    wb = Workbook()
    meta = wb.active
    meta.title = META_NAME
    meta.append(["__AMS_META__"])
    meta.append(["kind", ARCHIVE_KIND])
    meta.append(["format_version", FORMAT_VERSION])
    meta.append(["exported_at", datetime.utcnow().isoformat()])
    meta.append(["tables", len(names)])
    for c in meta[1]:
        c.font = Font(bold=True)

    total = 0
    for name in names:
        cols = [c["name"] for c in table_columns(conn, name)]
        ws = wb.create_sheet(_sheet_name(name))
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1F2A44")
            cell.font = Font(bold=True, color="FFFFFF")
        for tup in conn.execute(f'SELECT * FROM "{name}"'):
            ws.append([_cell(v) for v in tup])
            total += 1
    wb.save(dest)
    return {"path": str(dest), "tables": len(names), "rows": total}


def _cell(value):
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, float) and value != value:
        return None
    return value
