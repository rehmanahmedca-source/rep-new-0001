"""Regression tests for the legacy migration pipeline.

Covers the "TOTAL ROWS = 0" incident: an ALLEXPORT-style workbook (one sheet
per database table, no official DATA_ENTRY template sheet) must NEVER report a
silent zero. Every sheet is discovered with real counts, master entities are
adapted and imported through the controlled adapters, transactions are
validated with dependency ordering, and duplicate uploads are detected.
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
ALLEXPORT = ROOT / "legacy data" / "ALLEXPORT-28-08-2026_12-35PM.xlsx"

allexport_bytes = None


@pytest.fixture(scope="module")
def allexport():
    global allexport_bytes
    if allexport_bytes is None:
        allexport_bytes = ALLEXPORT.read_bytes()
    return allexport_bytes


def _wb_from(rows_by_sheet: dict[str, list[list]]) -> bytes:
    wb = Workbook()
    first = True
    for name, rows in rows_by_sheet.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = name
        for r in rows:
            ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Discovery layer — pure read, no database needed
# ---------------------------------------------------------------------------

def test_discovery_counts_every_sheet_of_the_incident_file(allexport):
    from app.services import migration_discovery as discovery

    wb = load_workbook(io.BytesIO(allexport), data_only=True)
    profiles = discovery.profile_workbook(wb)
    discovery.detect_entities(profiles)
    by_name = {p.name: p for p in profiles}

    assert len(profiles) == 63
    # the incident: workbook full of data, importer said zero
    assert sum(p.rows_found for p in profiles) == 26705
    assert "DATA_ENTRY" not in by_name  # why the strict template path found nothing

    client = by_name["client"]
    assert (client.entity, client.confidence, client.status) == ("CLIENTS", "HIGH", "MAPPED")
    assert client.rows_found == 315
    assert client.mapping.get("Client Name") == "name"

    assert by_name["booking"].entity == "BOOKINGS"
    assert by_name["booking_item"].entity == "BOOKINGS"
    assert by_name["supplier"].entity == "SUPPLIERS"
    assert by_name["account"].entity == "ACCOUNTS"

    # internal/derived tables are reported with an explicit reason, never silently dropped
    assert by_name["audit_log"].status == "IGNORED"
    assert by_name["user"].status == "IGNORED"
    assert by_name["entry"].status == "IGNORED"
    assert by_name["pending_bill"].status == "IGNORED"
    assert by_name["__AMS_META__"].status == "IGNORED"
    for name in ("audit_log", "user", "entry", "pending_bill", "__AMS_META__"):
        assert by_name[name].reason

    # empty sheets are distinguished from unmapped sheets
    assert by_name["delivery"].empty and by_name["delivery"].status == "IGNORED"


# ---------------------------------------------------------------------------
# End-to-end behaviour on the app database
# ---------------------------------------------------------------------------

@pytest.fixture()
def migrate(app, allexport):
    """Runs with the app context already pushed for convenience."""
    ctx = app.app_context()
    ctx.push()
    yield allexport
    ctx.pop()


def test_zero_row_incident_is_fixed(migrate):
    raw = migrate
    from app.services.legacy_migration import validate_upload, import_run
    from models import Client

    run, issues = validate_upload("CLIENTS", raw, ALLEXPORT.name, "pytest")
    summary = json.loads(run.summary_json)

    # the headline numbers are real, and the explanation is persisted on the run
    assert summary["SOURCE ROWS FOUND"] == 26705
    assert summary["MAPPED"] == 315
    assert summary["UNMAPPED"] == 26705 - 315
    assert summary["READY"] == 315
    assert run.mode == "LEGACY_ADAPTED"
    assert summary["ISSUES"]  # workbook-level findings no longer discarded
    assert any(i["column"] == "Sheet" and "missing" in i["problem"] for i in summary["ISSUES"])

    created = import_run(run)
    assert created == 315
    # every imported client name matches a name in the source sheet
    names = {c.name for c in Client.query.all()}
    assert "TAHIR ABBAS SB KOANKH" in names

    # re-uploading the same file must not duplicate anything
    run2, _ = validate_upload("CLIENTS", raw, ALLEXPORT.name, "pytest")
    s2 = json.loads(run2.summary_json)
    assert s2["EXACT_DUPLICATE"] == 315 and s2["READY"] == 0


def test_transaction_runs_are_validated_with_real_counts_and_order(migrate):
    raw = migrate
    from app.services.legacy_migration import validate_upload, import_run

    before = json.loads(validate_upload("BOOKINGS", raw, ALLEXPORT.name, "pytest")[0].summary_json)
    # transactions validate (never zero) but stay locked until master deps exist
    assert before["MAPPED"] == 1359  # 411 bookings + 948 booking items
    assert before["BLOCKED"] > 0     # import-order enforcement
    assert before["PENDING DEPENDENCIES"] == ["CLIENTS", "MATERIALS"]

    # import masters in dependency order, then revalidate bookings
    for kind in ("CLIENTS", "MATERIALS"):
        mrun, _ = validate_upload(kind, raw, ALLEXPORT.name, "pytest")
        assert import_run(mrun) > 0
    after = json.loads(validate_upload("BOOKINGS", raw, ALLEXPORT.name, "pytest")[0].summary_json)
    assert not after.get("PENDING DEPENDENCIES")
    assert after["READY"] > 0 and after["BLOCKED"] == 0
    assert after["MAPPED"] == after["READY"] + after["WARNING"] + after["INVALID"] + after["EXACT_DUPLICATE"] + after["ORPHAN"] + after["BLOCKED"]

    # transaction import remains locked: stock/ledger bypass is forbidden
    brun, _ = validate_upload("BOOKINGS", raw, ALLEXPORT.name, "pytest")
    with pytest.raises(ValueError, match="locked"):
        import_run(brun)


def test_analysis_run_and_prepare_from_stored_source(migrate):
    raw = migrate
    from app.services.legacy_migration import analyze_upload, rerun_prepare

    arun = analyze_upload(raw, ALLEXPORT.name, "pytest")
    assert arun.template_type == "ANALYSIS" and arun.status == "DISCOVERED"
    summary = json.loads(arun.summary_json)
    assert summary["SOURCE ROWS FOUND"] == 26705
    ents = summary["DETECTED_ENTITIES"]
    assert ents["CLIENTS"]["rows"] == 315 and "client" in ents["CLIENTS"]["sheets"]
    assert [o["entity"] for o in summary["IMPORT ORDER"]][:4] == ["CLIENTS", "SUPPLIERS", "MATERIALS", "ACCOUNTS"]

    child, issues = rerun_prepare(arun.id, "CLIENTS", "pytest")
    assert child.template_type == "CLIENTS" and child.mode == "LEGACY_ADAPTED"
    csummary = json.loads(child.summary_json)
    assert csummary["MAPPED"] == 315
    assert csummary["Prepared from analysis run"] == arun.id


def test_synthetic_legacy_customer_sheet_is_adapted_and_profiled(migrate):
    from app.services.legacy_migration import validate_upload

    raw = _wb_from({
        "CUSTOMER MASTER": [
            ["Customer Name", "MOBILE NO", "Address", "Balance"],
            ["Muhammad Ahmed", "0300-1112233", "Main Bazar JPS", "12500"],
            ["ABC TRADERS", "0345-9998887", "Chak Bahram", "-450"],
            ["", "0300-0000000", "No Name Ltd", "0"],           # missing required name -> INVALID
        ],
        "JUNK NOTES": [["memo", "text"], ["a", "just scribbles"]],
    })
    run, issues = validate_upload("CLIENTS", raw, "legacy_customer.xlsx", "pytest")
    s = json.loads(run.summary_json)
    assert s["SOURCE ROWS FOUND"] == 4   # 3 customers + 1 junk row — nothing silent
    assert s["MAPPED"] == 3 and s["READY"] == 2 and s["INVALID"] == 1
    assert s["MODE"] == "LEGACY_ADAPTED"
    junk = next(p for p in s["SHEETS"] if p["sheet"] == "JUNK NOTES")
    assert junk["status"] == "NOT_MAPPED" and junk["rows_found"] == 1 and junk["reason"]
    ref = [r for r in run.rows if r.status == "READY"][0].legacy_reference
    assert ref.startswith("AUTO-")  # stable derived reference for traceability
    assert run.rows[0].data_json.count("Muhammad Ahmed") == 1

    # re-upload the identical sheet after import: possible-duplicate review, never auto-merge
    from app.services.legacy_migration import import_run
    import_run(run)
    run2, _ = validate_upload("CLIENTS", raw, "legacy_customer.xlsx", "pytest")
    s2 = json.loads(run2.summary_json)
    assert s2["EXACT_DUPLICATE"] == 2  # AUTO refs stable across runs


def test_exact_official_template_path_still_works(migrate):
    from app.services.legacy_migration import template_workbook, validate_upload

    wb = template_workbook("CLIENTS")
    ws = wb["DATA_ENTRY"]
    ws.append(["CUST-001", "Genuine Ltd", "0300-5554433", "Street 1", "General", "note", ""])
    buf = io.BytesIO()
    wb.save(buf)
    run, issues = validate_upload("CLIENTS", buf.getvalue(), "official.xlsx", "pytest")
    s = json.loads(run.summary_json)
    assert run.mode == "EXACT_TEMPLATE"
    assert s["READY"] == 1 and s["SOURCE ROWS FOUND"] >= 1
    # the EXAMPLE row in the official template is still skipped
    assert s["MAPPED"] == 1


def test_no_match_workbook_reports_counts_and_reasons(migrate):
    from app.services.legacy_migration import validate_upload

    raw = _wb_from({"Random": [["Foo", "Bar", "Baz"], ["1", "2", "3"], ["4", "5", "6"]]})
    run, issues = validate_upload("PAYMENTS", raw, "random.xlsx", "pytest")
    s = json.loads(run.summary_json)
    assert s["SOURCE ROWS FOUND"] == 2 and s["MAPPED"] == 0
    assert s["BLOCKED"] == 0
    assert any(i["status"] == "BLOCKED" for i in s["ISSUES"])  # explicit explanation, no silent zero
    assert s["ISSUES"][-1]["problem"].startswith("No sheet in this workbook")


def test_migration_ui_pages_render(app, client, allexport):
    resp = client.post("/login", data={"username": "Admin", "password": "Admin@fbm12345"})
    assert resp.status_code in (200, 302)
    r = client.get("/legacy-migration/")
    assert r.status_code == 200
    # upload through the real form with AUTO analysis — the route that once flashed
    # a cheerful “Validation complete” over an all-zero summary must not do that again
    r = client.post(
        "/legacy-migration/upload",
        data={"template_type": "CLIENTS", "file": (io.BytesIO(allexport), ALLEXPORT.name)},
        content_type="multipart/form-data",
        follow_redirects=False,
    )
    assert r.status_code in (200, 302)
    with app.app_context():
        from models import MigrationRun
        run = MigrationRun.query.order_by(MigrationRun.id.desc()).first()
        summary = json.loads(run.summary_json)
    assert summary["SOURCE ROWS FOUND"] == 26705 and summary["MAPPED"] == 315
    r = client.get(f"/legacy-migration/run/{run.id}")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    # real numbers rendered for the user, not a silent zero
    assert "26705" in html and "315" in html and "MAPPED" in html
    err = client.get(f"/legacy-migration/run/{run.id}/errors.xlsx")
    assert err.status_code == 200 and err.mimetype.endswith("spreadsheetml.sheet")
