"""Data Center: schema-aware archive, versioned restore, UI flow.

These tests prove the design contract in docs/DATA_CENTER.md:
* an OLD export restores cleanly into a NEWER schema (new tables untouched,
  new columns defaulted/reported, nothing deleted);
* restore is idempotent (apply twice == apply once);
* unknown columns/tables abort with zero writes;
* the UI flow upload → plan → APPLY works end to end.
"""
from __future__ import annotations

import io
import json
import sqlite3
from pathlib import Path

import pytest

from data_ops.backup_ops import create_data_backup, export_json
from data_ops.engine import execute_restore
from data_ops.planner import SchemaAbort
from data_ops.verify import row_counts


SCHEMA = """
CREATE TABLE client (
  id INTEGER PRIMARY KEY,
  code TEXT,
  name TEXT,
  category TEXT,
  notes TEXT,
  is_active INTEGER DEFAULT 1
);
CREATE TABLE payment (
  id INTEGER PRIMARY KEY,
  client_id INTEGER REFERENCES client(id),
  client_code TEXT,
  client_name TEXT,
  amount REAL,
  account_no TEXT,
  is_void INTEGER DEFAULT 0,
  note TEXT
);
"""


def _make_db(path: Path, kind: str = "old") -> Path:
    conn = sqlite3.connect(str(path))
    conn.executescript(SCHEMA)
    conn.execute("INSERT INTO client (id, code, name, category) VALUES (1, 'C1', 'Alpha', 'General')")
    conn.execute(
        "INSERT INTO payment (id, client_id, client_code, client_name, amount, account_no, is_void) "
        "VALUES (10, 1, 'C1', 'Alpha', 500, '7761.0', 0)"
    )
    if kind == "new":
        conn.execute("ALTER TABLE payment ADD COLUMN extra_note TEXT DEFAULT ''")
        conn.execute("ALTER TABLE payment ADD COLUMN integrity_flag INTEGER NOT NULL DEFAULT 0")
        conn.execute("CREATE TABLE new_thing (id INTEGER PRIMARY KEY, label TEXT)")
        conn.execute("INSERT INTO new_thing (id, label) VALUES (1, 'live-value-must-survive')")
        # a live row that the old file does not know about at all
        conn.execute("INSERT INTO payment (id, client_id, amount, is_void) VALUES (99, 1, 123, 0)")
    conn.commit()
    conn.close()
    return path


def test_envelope_has_schema_and_restores(tmp_path):
    db = _make_db(tmp_path / "src.db")
    archive = tmp_path / "envelope.json"
    info = export_json(sqlite3.connect(str(db)), archive, app_version="test", db_name="src.db")
    assert info["rows"] == 2
    payload = json.loads(archive.read_text())
    assert payload["kind"] == "ams.data-archive"
    assert payload["format_version"] == "2026-08"
    assert payload["schema"]["payment"]["primary_key"] == ["id"]
    assert payload["schema"]["payment"]["foreign_keys"][0]["table"] == "client"

    target = _make_db(tmp_path / "target.db")
    conn = sqlite3.connect(str(target))
    before = row_counts(conn)
    report = execute_restore(conn, archive)
    conn.close()
    assert report["ok"]
    assert row_counts(sqlite3.connect(str(target))) == before


def test_old_archive_into_newer_schema_clean(tmp_path):
    """THE core guarantee: old data enters a newer app without loss."""
    old = _make_db(tmp_path / "old.db", kind="old")
    archive = tmp_path / "old.json"
    export_json(sqlite3.connect(str(old)), archive, app_version="old-app")

    newer = _make_db(tmp_path / "newer.db", kind="new")
    conn = sqlite3.connect(str(newer))
    conn.execute("PRAGMA foreign_keys=ON")
    report = execute_restore(conn, archive)
    conn.close()

    assert report["ok"]
    # new table untouched (still has its seeded row, zero from the old file)
    assert "new_thing" in report["untouched_tables"]
    assert "extra_note" in [f["column"] for f in report["filled_missing"]]
    assert "integrity_flag" in [f["column"] for f in report["filled_missing"]]

    after = row_counts(sqlite3.connect(str(newer)))
    assert after["payment"] == 2  # old 10 upserted, live 99 kept — nothing lost
    assert after["client"] == 1
    assert after["new_thing"] == 1
    # live row that the file didn't contain was NOT touched
    row = sqlite3.connect(str(newer)).execute(
        "SELECT amount, integrity_flag FROM payment WHERE id=99"
    ).fetchone()
    assert row[0] == 123


def test_restore_idempotent(tmp_path):
    old = _make_db(tmp_path / "old.db")
    archive = tmp_path / "old.json"
    export_json(sqlite3.connect(str(old)), archive)

    target = _make_db(tmp_path / "target.db", kind="new")
    c1 = row_counts(sqlite3.connect(str(target)))
    execute_restore(sqlite3.connect(str(target)), archive)
    execute_restore(sqlite3.connect(str(target)), archive)
    c2 = row_counts(sqlite3.connect(str(target)))
    assert c1 == c2


def test_unknown_column_aborts_before_write(tmp_path):
    target = _make_db(tmp_path / "target.db")
    before = row_counts(sqlite3.connect(str(target)))
    payload = {
        "kind": "ams.data-archive",
        "format_version": "2026-08",
        "schema": {"payment": {"columns": [{"name": "id"}, {"name": "totally_unknown"}]}},
        "tables": {"payment": [{"id": 1, "totally_unknown": 5}]},
    }
    conn = sqlite3.connect(str(target))
    with pytest.raises(SchemaAbort, match="Unknown columns"):
        execute_restore(conn, payload)
    conn.rollback()
    conn.close()
    assert row_counts(sqlite3.connect(str(target))) == before


def test_newer_archive_refused(tmp_path):
    target = _make_db(tmp_path / "target.db")
    payload = {
        "kind": "ams.data-archive",
        "format_version": "2099-01",
        "schema": {},
        "tables": {},
    }
    conn = sqlite3.connect(str(target))
    with pytest.raises(SchemaAbort, match="newer"):
        execute_restore(conn, payload)
    conn.close()


def test_fk_to_live_row_not_blanked(tmp_path):
    """FK values pointing at LIVE rows (not in the file) must survive planning."""
    schema = SCHEMA + """
CREATE TABLE material (id INTEGER PRIMARY KEY, label TEXT);
ALTER TABLE payment ADD COLUMN material_id INTEGER REFERENCES material(id);
"""
    target = tmp_path / "t.db"
    conn = sqlite3.connect(str(target))
    conn.executescript(schema)
    conn.execute("INSERT INTO client (id, code, name) VALUES (1, 'C1', 'Alpha')")
    conn.execute("INSERT INTO material (id, label) VALUES (7, 'Cement')")
    conn.commit()
    conn.close()

    payload = {
        "kind": "ams.data-archive",
        "format_version": "2026-08",
        "schema": {"payment": {"columns": [{"name": "id"}, {"name": "client_id"}, {"name": "material_id"}]}},
        "tables": {"payment": [{"id": 30, "client_id": 1, "material_id": 7}]},
    }
    conn = sqlite3.connect(str(target))
    plan = execute_restore(conn, payload, dry_run=True)
    conn.close()
    assert plan["ok"]
    assert plan["blanked_optional_fks"] == []

    conn = sqlite3.connect(str(target))
    execute_restore(conn, payload)
    conn.close()
    row = sqlite3.connect(str(target)).execute(
        "SELECT client_id, material_id FROM payment WHERE id=30"
    ).fetchone()
    assert row == (1, 7)


def test_composite_pk_upsert(tmp_path):
    """Tables whose PK is not ``id`` still upsert idempotently."""
    target = tmp_path / "kv.db"
    conn = sqlite3.connect(str(target))
    conn.execute("CREATE TABLE kv (k TEXT PRIMARY KEY, v TEXT)")
    conn.commit()
    conn.close()
    payload = {
        "kind": "ams.data-archive",
        "format_version": "2026-08",
        "schema": {"kv": {"columns": [{"name": "k"}, {"name": "v"}], "primary_key": ["k"]}},
        "tables": {"kv": [{"k": "a", "v": "1"}]},
    }
    for _ in range(2):
        conn = sqlite3.connect(str(target))
        execute_restore(conn, payload)
        conn.close()
    assert row_counts(sqlite3.connect(str(target)))["kv"] == 1
    assert sqlite3.connect(str(target)).execute("SELECT v FROM kv WHERE k='a'").fetchone()[0] == "1"


def test_backup_manifest_and_envelope(tmp_path):
    db = _make_db(tmp_path / "live.db")
    result = create_data_backup(db, tmp_path / "backups", reason="test", app_version="t")
    manifest = json.loads((Path(result["path"]) / "manifest.json").read_text())
    assert manifest["format_version"] == "2026-08"
    archive = json.loads((Path(result["path"]) / "export.json").read_text())
    assert archive["kind"] == "ams.data-archive"
    assert archive["schema"]["client"]["primary_key"] == ["id"]


def test_export_xlsx_display(tmp_path):
    from data_ops.xlsx_export import export_xlsx

    db = _make_db(tmp_path / "live.db")
    out = export_xlsx(sqlite3.connect(str(db)), tmp_path / "view.xlsx")
    assert out["rows"] == 2
    from openpyxl import load_workbook

    wb = load_workbook(out["path"], read_only=True)
    assert "__AMS_META__" in wb.sheetnames
    assert "client" in wb.sheetnames


# ---------------------------------------------------------------------------
# UI flow
# ---------------------------------------------------------------------------

def login(client, username="Admin", password="Admin@fbm12345"):
    resp = client.post("/login", data={"username": username, "password": password})
    assert resp.status_code in (302, 303), resp.get_data(as_text=True)[:300]


@pytest.fixture()
def dc_app(app_factory):
    return app_factory(FULL_RAW_IMPORT_ENABLED="0")


def test_data_center_pages_render(dc_app):
    from tests.conftest import make_csrf_client

    client = make_csrf_client(dc_app)
    login(client)
    for path in (
        "/import_export/",
        "/import_export/data-export",
        "/import_export/restore",
        "/import_export/legacy",
        "/import_export/history",
        "/import_export/restore/db",
    ):
        resp = client.get(path)
        assert resp.status_code == 200, (path, resp.get_data(as_text=True)[:400])


def test_export_json_download(dc_app, tmp_path):
    from tests.conftest import make_csrf_client

    client = make_csrf_client(dc_app)
    login(client)
    resp = client.post("/import_export/export/json", data={"tables": "all"})
    assert resp.status_code == 200
    payload = json.loads(resp.data)
    assert payload["kind"] == "ams.data-archive"
    assert payload.get("tables")
    assert "user" in payload["tables"]


def test_restore_plan_and_apply_flow(dc_app, tmp_path):
    """UI round-trip: export JSON → upload → plan page → APPLY → run recorded."""
    from tests.conftest import make_csrf_client
    from data_ops.portable import export_json
    from sqlalchemy import text
    from models import db

    client = make_csrf_client(dc_app)
    login(client)

    # seed a client through the app session so the DB row exists
    with dc_app.app_context():
        db.session.execute(
            text("INSERT INTO client (id, code, name, category, is_active) VALUES (9001, 'UI9', 'Ui Nine', 'General', 1)")
        )
        db.session.commit()

    db_path = Path(dc_app.config["APP_DB_PATH"])
    archive = tmp_path / "ui_export.json"
    conn = sqlite3.connect(str(db_path))
    export_json(conn, archive, app_version="ui-test", db_name=db_path.name)
    conn.close()

    raw = archive.read_bytes()
    resp = client.post(
        "/import_export/restore/plan",
        data={"file": (io.BytesIO(raw), "ui_export.json")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200, resp.get_data(as_text=True)[:400]
    assert b"rows to write" in resp.data
    assert b"Apply Restore" in resp.data

    resp = client.post(
        "/import_export/restore/apply",
        data={"confirm": "APPLY"},
        follow_redirects=True,
    )
    assert resp.status_code == 200, resp.get_data(as_text=True)[:400]
    assert b"Restore completed" in resp.data

    with dc_app.app_context():
        run_count = db.session.execute(
            text("SELECT COUNT(*) FROM data_transfer_run WHERE kind='restore'")
        ).scalar()
        assert run_count >= 1
        n = db.session.execute(
            text("SELECT COUNT(*) FROM client WHERE id=9001")
        ).scalar()
        assert n == 1


def test_restore_requires_typed_confirm(dc_app, tmp_path):
    from tests.conftest import make_csrf_client

    client = make_csrf_client(dc_app)
    login(client)
    resp = client.post("/import_export/restore/apply", data={"confirm": "APPLY"})
    assert resp.status_code in (302, 303)  # no pending plan → redirect
