"""Controlled ETL helpers. Templates carry facts; application state is calculated later.

Pipeline (post-fix): the importer FIRST profiles every sheet of the uploaded
workbook (discovery layer), THEN decides per sheet whether it is an exact
official template sheet, a legacy sheet that can be safely adapted into the
template shape, or a sheet that must be reported as NOT MAPPED / IGNORED with
an explicit reason and its real row count. A "VALIDATED with 0 rows" summary is
no longer possible for a non-empty workbook: SOURCE ROWS FOUND always reflects
the real contents, and workbook-level issues are persisted on the run — before
this change they were computed, returned to the route, and thrown away, which
is how "0 rows everywhere" became indistinguishable from a silently rejected
file.
"""
from __future__ import annotations
import hashlib, io, json, re, uuid
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from flask import current_app
from sqlalchemy import func
from models import db, Client, Supplier, Material, MaterialCategory, Account, MigrationRun, MigrationRow, MigrationMapping
from app.services import migration_discovery as discovery

# This matrix is deliberately tied to current SQLAlchemy models, not an old XLSX schema.
# Legacy workbooks are ADAPTED into these shapes by app.services.migration_discovery;
# they are no longer required to match them byte-for-byte.
TEMPLATES = {
 'CLIENTS': {'file':'01_Clients_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Client Name*','Phone','Address','Category','Notes','Legacy Expected Due']}, 'dependency':[]},
 'SUPPLIERS': {'file':'02_Suppliers_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Supplier Name*','Phone','Address','Notes','Legacy Expected Due']}, 'dependency':[]},
 'MATERIALS': {'file':'03_Materials_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Material Name*','Category','Unit','Unit Price','Notes','Legacy Expected Stock']}, 'dependency':[]},
 'ACCOUNTS': {'file':'04_Accounts_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Account Name*','Category*','Account Type*','Opening Balance','Bank Name','Account Number','Notes','Legacy Expected Balance']}, 'dependency':[]},
 'GRN': {'file':'05_GRN_Migration_Template.xlsx','sheets': {'GRN_HEADERS':['Legacy GRN Reference*','GRN Number','Date*','Supplier*','Account','Notes'], 'GRN_ITEMS':['Legacy GRN Reference*','Material*','Quantity*','Rate','Discount','Notes']}, 'dependency':['SUPPLIERS','MATERIALS']},
 'BOOKINGS': {'file':'06_Bookings_Migration_Template.xlsx','sheets': {'BOOKINGS':['Legacy Booking Reference*','Booking Number','Date*','Client*','Notes'], 'BOOKING_ITEMS':['Legacy Booking Reference*','Material*','Quantity*','Rate','Discount']}, 'dependency':['CLIENTS','MATERIALS']},
 'SALES': {'file':'07_Sales_Migration_Template.xlsx','sheets': {'SALES':['Legacy Sale Reference*','Sale/Bill Number','Date*','Client*','Legacy Booking Reference','Sale Type*','Account','Notes'], 'SALE_ITEMS':['Legacy Sale Reference*','Material*','Quantity*','Rate','Discount']}, 'dependency':['CLIENTS','MATERIALS']},
 'DIRECT_SALES': {'file':'08_Direct_Sales_Migration_Template.xlsx','sheets': {'SALES':['Legacy Sale Reference*','Bill Number','Date*','Client / Walk-in Name','Account','Payment Type','Notes'], 'SALE_ITEMS':['Legacy Sale Reference*','Material*','Quantity*','Rate','Discount']}, 'dependency':['MATERIALS']},
 'DELIVERIES': {'file':'09_Deliveries_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Delivery Reference*','Date*','Client','Bill Number','Material*','Quantity*','Notes']}, 'dependency':['MATERIALS']},
 'PAYMENTS': {'file':'10_Payments_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Payment Reference*','Date*','Party Type*','Party*','Amount*','Account*','Payment Type*','Reference Number','Notes']}, 'dependency':['ACCOUNTS']},
 'EXPENSES': {'file':'11_Expenses_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Expense Reference*','Date*','Account*','Amount*','Category','Reference Number','Notes']}, 'dependency':['ACCOUNTS']},
 'OPENING_BALANCES': {'file':'12_Opening_Balances_Migration_Template.xlsx','sheets': {'DATA_ENTRY':['Legacy Reference*','Balance Type*','Party / Account / Material*','Amount or Quantity*','Date*','Notes','Legacy Expected Balance']}, 'dependency':[]},
}

MASTER_KINDS = {'CLIENTS','SUPPLIERS','MATERIALS','ACCOUNTS'}
REFERENCE_FIELDS = {'Client':'CLIENTS','Supplier':'SUPPLIERS','Party':'BOTH','Material':'MATERIALS','Account':'ACCOUNTS'}
REFERENCE_LOOKUP = {'CLIENTS': Client, 'SUPPLIERS': Supplier, 'MATERIALS': Material, 'ACCOUNTS': Account}
NAME_FIELDS = {'CLIENTS':'Client Name','SUPPLIERS':'Supplier Name','MATERIALS':'Material Name','ACCOUNTS':'Account Name'}
REFERENCE_COLUMNS = ('Legacy Reference','Legacy GRN Reference','Legacy Booking Reference','Legacy Sale Reference','Legacy Payment Reference','Legacy Delivery Reference','Legacy Expense Reference')
NUMERIC_CHECK_FIELDS = ('Date','Quantity','Rate','Amount','Opening Balance','Amount or Quantity','Unit Price','Discount','Legacy Expected Due','Legacy Expected Stock','Legacy Expected Balance')

def norm(value): return re.sub(r'\s+', ' ', str(value or '').strip()).casefold()
def clean(value): return '' if value is None else str(value).strip()
def rowhash(data): return hashlib.sha256(json.dumps(data, sort_keys=True, default=str).encode()).hexdigest()
def filehash(raw): return hashlib.sha256(raw).hexdigest()
def required(header): return header.endswith('*')
def bare(header): return header.rstrip('*').strip()

def _reference_value(data):
    for cand in REFERENCE_COLUMNS:
        if data.get(cand): return str(data[cand]).strip()
    return None

def _uploads_dir() -> Path:
    path = Path(current_app.instance_path) / 'migration' / 'uploads'
    path.mkdir(parents=True, exist_ok=True)
    return path

def _save_source(raw: bytes, run_key: str) -> str | None:
    """Preserve the original legacy source for review/replay (never modified)."""
    try:
        target = _uploads_dir() / f'{run_key}.xlsx'
        target.write_bytes(raw)
        return str(target.relative_to(Path(current_app.instance_path)))
    except OSError:
        return None

def _load_source(run) -> bytes | None:
    summary = json.loads(run.summary_json or '{}')
    rel = summary.get('STORED_SOURCE')
    if not rel: return None
    path = Path(current_app.instance_path) / rel
    return path.read_bytes() if path.exists() else None


def template_workbook(kind):
    spec=TEMPLATES[kind]; wb=Workbook(); ins=wb.active; ins.title='INSTRUCTIONS'
    ins.append(['LEGACY DATA MIGRATION — '+kind]); ins.append(['Purpose','Enter cleaned historical source facts only. Calculated balances, stock and profits never overwrite this application.'])
    ins.append(['How to use','Fill only the named data sheets. Keep the header row unchanged. Use YYYY-MM-DD dates and plain numbers.'])
    ins.append(['Duplicate rule','Legacy Reference is mandatory and is permanent source traceability. Re-uploading identical rows is skipped.'])
    ins.append(['Common mistake','Do not put multiple materials or parties in one cell; use one item per row on item sheets.'])
    ins.append(['Required fields','Columns marked * are required. Resolve suggested matches yourself; the system never silently selects one.'])
    ins.append(['Legacy workbooks','Any workbook can be uploaded through AUTO analysis: sheets are discovered, entities detected, columns mapped. Only HIGH/MEDIUM confidence sheets are adapted; medium-confidence rows require review before import.'])
    ins.column_dimensions['A'].width=26; ins.column_dimensions['B'].width=115
    for sheet, headers in spec['sheets'].items():
        ws=wb.create_sheet(sheet); ws.append(headers); ws.append(['EXAMPLE-'+kind[:3]+'-001'] + ['Example value']*(len(headers)-1)); ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78'); ws.column_dimensions[c.column_letter].width=max(16,min(32,len(c.value)+8))
    # Composite transaction workbooks keep their named header/item sheets; a small DATA_ENTRY index preserves the standard workbook shape.
    if 'DATA_ENTRY' not in wb.sheetnames:
        index=wb.create_sheet('DATA_ENTRY'); index.append(['Use the dedicated sheets: '+', '.join(spec['sheets'].keys())]); index.append(['Do not enter data on this index sheet.'])
    ex=wb.create_sheet('EXAMPLES'); ex.append(['Read INSTRUCTIONS first. Sample rows are illustrative and must be replaced.'])
    ref=wb.create_sheet('REFERENCE_DATA'); ref.append(['Current valid values are a guide only; they never modify uploaded facts.'])
    for title, values in [('Clients',[x.name for x in Client.query.order_by(Client.name).all()]),('Suppliers',[x.name for x in Supplier.query.order_by(Supplier.name).all()]),('Materials',[x.name for x in Material.query.order_by(Material.name).all()]),('Accounts',[x.name for x in Account.query.order_by(Account.name).all()])]:
        ref.append([title]); [ref.append([v]) for v in values]
    return wb

def _problem(sheet,row,col,message,status='INVALID',suggestions=None):
    return {'sheet':sheet,'row':row,'column':col,'problem':message,'suggested_action':suggestions or '', 'status':status}

# ---------------------------------------------------------------------------
# Row extraction
# ---------------------------------------------------------------------------

def _exact_headers(ws, headers):
    actual=[clean(c.value) for c in ws[1]]
    while actual and not actual[-1]: actual.pop()
    return actual == headers

def _extract_exact(ws, headers):
    """Rows from a sheet whose headers match the official template exactly."""
    out=[]
    for excel_row, cells in enumerate(ws.iter_rows(min_row=2, values_only=True),2):
        data={bare(headers[i]): clean(cells[i] if i<len(cells) else '') for i in range(len(headers))}
        ref=_reference_value(data)
        if not any(data.values()) or (ref and ref.startswith('EXAMPLE-')): continue
        out.append((excel_row, data, []))
    return out

# ---------------------------------------------------------------------------
# Row validation (shared by exact-template and adapted-legacy rows)
# ---------------------------------------------------------------------------

def _profile_row(sheet, excel_row, data, kind, notes=None):
    errors=[]
    headers = TEMPLATES[kind]['sheets'][sheet]
    for h in headers:
        if required(h) and not str(data.get(bare(h)) or '').strip():
            errors.append(_problem(sheet,excel_row,bare(h),'This required value is missing.','INVALID','Fill the value in the source, or extend the column mapping so it is derived.'))
    for key in NUMERIC_CHECK_FIELDS:
        v=str(data.get(key) or '').strip()
        if v:
            try:
                if key=='Date': datetime.fromisoformat(v.replace('Z','+00:00').strip())
                else: float(v.replace(',',''))
            except ValueError:
                errors.append(_problem(sheet,excel_row,key,'Use YYYY-MM-DD for dates or a plain numeric value.','INVALID','Fix the cell value in the source workbook.'))
    for note in notes or []:
        # Policy class B: required-but-defaultable fields use a controlled
        # default and are recorded — informational, does not block the row.
        errors.append(_problem(sheet,excel_row,note.get('column',''),'Missing field resolved automatically: '+note.get('message',''),'NOTE','Review derived values before import.'))
    return errors

def _check_business_duplicates(kind, data, excel_row, sheet):
    """Name-collision protection for master rows: never auto-merge, always review."""
    if kind not in MASTER_KINDS: return []
    name=str(data.get(NAME_FIELDS[kind],'')).strip()
    if not name: return []
    model=REFERENCE_LOOKUP[kind]
    matches=model.query.filter(func.lower(model.name)==norm(name)).limit(4).all()
    if not matches: return []
    listing=', '.join(f'#{m.id} {m.name}' for m in matches)
    return [_problem(sheet,excel_row,NAME_FIELDS[kind],f'Name already exists in the ERP: {listing}. Flagged as possible duplicate — not merged automatically.','WARNING',f'Confirm merge vs new record before importing. Candidates: {listing}')]

def _check_references(kind, data, excel_row, sheet):
    """Reference resolution for party columns: names must resolve to a record;
    unresolved ones become ORPHAN rows with candidate suggestions (the orphan
    resolution queue the policy demands — nothing fails silently)."""
    issues=[]
    headers=TEMPLATES[kind]['sheets'][sheet]
    bare_map={bare(h):h for h in headers}
    for field, entity in REFERENCE_FIELDS.items():
        if field not in bare_map: continue
        value=str(data.get(field,'')).strip()
        if not value: continue
        kinds_to_check=[entity] if entity!='BOTH' else ['CLIENTS','SUPPLIERS']
        found=False; candidates=[]
        for k in kinds_to_check:
            model=REFERENCE_LOOKUP[k]
            if model.query.filter(func.lower(model.name)==norm(value)).first():
                found=True; break
            candidates+=[f'{k[:-1]} #{h.id} {h.name}' for h in model.query.filter(model.name.ilike(f'%{value[:24]}%')).limit(3).all()]
        if not found:
            label='client/supplier' if entity=='BOTH' else entity[:-1].lower()
            hint='Orphan resolution queue: map to an existing record, create the missing master record, skip, or block — then revalidate.'
            if candidates: hint+=' Possible matches: '+'; '.join(candidates)
            issues.append(_problem(sheet,excel_row,field,f'Unresolved reference "{value}" — no matching {label} record exists.','ORPHAN',hint))
    return issues

def _dependency_block(kind):
    """Master templates that must be completed before this transaction kind can
    resolve references (dependency-aware import order)."""
    if kind not in TEMPLATES or not TEMPLATES[kind].get('dependency'): return []
    pending=[]
    for dep in TEMPLATES[kind]['dependency']:
        if dep=='CATEGORIES': continue
        done=MigrationRun.query.filter_by(template_type=dep,status='COMPLETED').first() or MigrationMapping.query.filter_by(template_type=dep).first()
        if not done: pending.append(dep)
    return pending

def _classify_row(errors):
    statuses=[e['status'] for e in errors]
    for s in ('INVALID','EXACT_DUPLICATE','BLOCKED','ORPHAN','WARNING'):
        if s in statuses: return s
    return 'READY'

# ---------------------------------------------------------------------------
# Run persistence — never a silent zero
# ---------------------------------------------------------------------------

def _persist_run(kind, raw, filename, actor, sheet_rows, issues, profiles, mode, extra_summary=None):
    """Create MigrationRun + MigrationRow records. sheet_rows entries:
    (template_sheet, excel_row, data, errors, source_sheet_name|None)."""
    run=MigrationRun(run_key=str(uuid.uuid4()),template_type=kind,filename=filename,source_hash=filehash(raw),status='VALIDATED',summary_json='{}',uploaded_by=actor,mode=mode)
    db.session.add(run); db.session.flush()
    counts={'READY':0,'INVALID':0,'EXACT_DUPLICATE':0,'WARNING':0,'ORPHAN':0,'BLOCKED':0}
    seen_hashes={}; refs={}
    parent_refs=set()  # legacy references seen on header sheets (for *_ITEMS orphans)
    for sheet,_r,data,_e,_s in sheet_rows:
        if not sheet.endswith('_ITEMS') and (r:=_reference_value(data)): parent_refs.add(r)
    pending_deps=_dependency_block(kind)
    for sheet,excel_row,data,errors,source_sheet in sheet_rows:
        ref=_reference_value(data)
        if sheet.endswith('_ITEMS') and ref and ref not in parent_refs:
            errors=errors+[_problem(sheet,excel_row,'Legacy Reference',f'Item row references parent "{ref}" which is not present among the header rows of this workbook.','ORPHAN','Provide the parent in the same file or map the reference; then revalidate.')]
        dup_prev=seen_hashes.get((sheet,rowhash(data)))
        if dup_prev:
            # Same content, possibly different source rows: a business-duplicate
            # candidate for human review — never auto-discarded (policy) and
            # never auto-imported (only READY rows import).
            errors=errors+[_problem(sheet,excel_row,'Row',f'Identical to source row {dup_prev} in this workbook — possible business duplicate.','WARNING','Review both rows; keep only the valid one, or accept both if they are genuinely separate records.')]
        if ref:
            key=(sheet,norm(ref))
            if key in refs:
                errors=errors+[_problem(sheet,excel_row,'Legacy Reference',f'Reference already used at source row {refs[key]} of this workbook.','EXACT_DUPLICATE','Merge or renumber references in the source file.')]
            else: refs[key]=excel_row
            existing=MigrationMapping.query.filter_by(template_type=kind,legacy_reference=ref).first()
            if existing: errors=errors+[_problem(sheet,excel_row,'Legacy Reference','This reference was already imported as record '+str(existing.entity_id),'EXACT_DUPLICATE','Already imported; row will be skipped.')]
        if pending_deps and any(e['status']=='ORPHAN' for e in errors):
            # A missing reference here is an ordering problem, not a data problem:
            # the master template that owns it has not been imported yet.
            errors=[e for e in errors if e['status']!='ORPHAN']+[
                _problem(sheet,excel_row,'Import order',f'Reference unresolvable until master template(s) {", ".join(pending_deps)} are imported; row kept BLOCKED for revalidation.','BLOCKED',f'Expected import order: {", ".join(pending_deps)} → {kind}.')]
        status=_classify_row(errors)
        counts[status]=counts.get(status,0)+1
        seen_hashes[(sheet,rowhash(data))]=dup_prev or excel_row
        db.session.add(MigrationRow(run_id=run.id,source_sheet=source_sheet or sheet,source_row=excel_row,legacy_reference=ref,row_hash=rowhash(data),status=status,data_json=json.dumps(data,default=str),error_json=json.dumps(errors)))
    total_found=sum(p.rows_found for p in profiles)
    recognized=sum(p.rows_found for p in profiles if p.confidence in ('HIGH','MEDIUM','LOW'))
    mapped_rows=len(sheet_rows)
    summary=dict(counts)
    summary['Total Rows']=mapped_rows
    summary['SOURCE ROWS FOUND']=total_found
    summary['RECOGNIZED']=recognized
    summary['MAPPED']=mapped_rows
    summary['UNMAPPED']=max(0,total_found-mapped_rows)
    summary['SHEETS']=[p.as_dict() for p in profiles]
    summary['ISSUES']=issues[:400]
    summary['MODE']=mode
    if kind in TEMPLATES: summary['IMPORT ENABLED']=kind in MASTER_KINDS
    if pending_deps: summary['PENDING DEPENDENCIES']=pending_deps
    if extra_summary: summary.update(extra_summary)
    run.summary_json=json.dumps(summary,default=str)
    db.session.commit()
    return run

def spec_header_sheets(kind):
    return [s for s in TEMPLATES[kind]['sheets'] if not s.endswith('_ITEMS')]

def validate_upload(kind, raw, filename, actor=''):
    """Validate an uploaded workbook against a template.

    Exact official-template sheets are parsed unchanged (legacy behaviour
    preserved); otherwise the discovery layer looks for legacy sheets that
    confidently represent the selected entity and adapts them through the
    column mapping. Every sheet is accounted for in the run summary — real row
    count, detected entity, confidence, and an explicit reason whenever a
    sheet is not imported.
    """
    if kind not in TEMPLATES: raise ValueError('Unknown official migration template.')
    try: wb=load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc: raise ValueError('The file is not a readable XLSX workbook.') from exc
    spec=TEMPLATES[kind]; issues=[]
    profiles=discovery.profile_workbook(wb)
    discovery.detect_entities(profiles)
    sheet_rows=[]; mode='EXACT_TEMPLATE'; matched_sheets=set()
    for sheet, headers in spec['sheets'].items():
        ws=wb[sheet] if sheet in wb.sheetnames else None
        if ws is None:
            issues.append(_problem(sheet,0,'Sheet','Required sheet is missing from the uploaded workbook.','INVALID','Use the official template, or upload through AUTO analysis so legacy sheets are discovered and mapped instead.')); continue
        if not _exact_headers(ws, headers):
            issues.append(_problem(sheet,1,'Header','Headers do not match the official template exactly; the sheet was not treated as an official template sheet.','INVALID','Correct the header row, or let the discovery layer adapt the legacy columns.')); continue
        matched_sheets.add(sheet)
        for excel_row,data,notes in _extract_exact(ws,headers):
            errors=_profile_row(sheet,excel_row,data,kind,notes)
            errors+=_check_business_duplicates(kind,data,excel_row,sheet)
            errors+=_check_references(kind,data,excel_row,sheet)
            sheet_rows.append((sheet,excel_row,data,errors,None))

    adapted_info={}
    if not matched_sheets:
        # ----------------------------- LEGACY ADAPTATION PATH -----------------------------
        all_header_candidates=[p for p in profiles if p.entity==kind and p.status=='MAPPED' and not discovery._is_item_sheet(p)]
        header_candidates=sorted([p for p in all_header_candidates if p.entity==kind and p.status=='MAPPED' and not discovery._is_item_sheet(p)], key=lambda p:-p.score)[:1]
        item_candidates=sorted([p for p in profiles if p.entity==kind and p.status=='MAPPED' and discovery._is_item_sheet(p)], key=lambda p:-p.score)[:1]
        unused=[p for p in all_header_candidates if p not in header_candidates]
        if header_candidates:
            for p in unused:
                issues.append(_problem(p.name,0,'Mapping',f"Sheet also resembles {kind} (score {p.score}) but '{header_candidates[0].name}' is the stronger match; only the best sheet is adapted automatically.",'WARNING','If this sheet is the right source, clean the workbook or request a custom mapping.'))
            mode='LEGACY_ADAPTED'
            id_ref_maps=discovery.internal_lookup_maps(profiles, lambda prof: wb[prof.name].iter_rows(values_only=True))
            rows_provider=lambda prof: wb[prof.name].iter_rows(values_only=True)
            target_header=list(spec['sheets'])[0]
            for p in header_candidates:
                p.target_sheet=target_header
                for excel_row,data,notes in discovery.adapt_sheet_rows(kind, p, rows_provider, id_ref_maps, target_sheet=target_header):
                    errors=_profile_row(target_header,excel_row,data,kind,notes)
                    if p.confidence=='MEDIUM':
                        errors.append(_problem(target_header,excel_row,'Mapping',f"Sheet '{p.name}' was matched to {kind} with MEDIUM confidence — row kept out of automatic import until reviewed.",'WARNING','Confirm the detected entity and mapping, then revalidate.'))
                    errors+=_check_business_duplicates(kind,data,excel_row,target_header)
                    errors+=_check_references(kind,data,excel_row,target_header)
                    sheet_rows.append((target_header,excel_row,data,errors,p.name))
                adapted_info[target_header]={'source_sheet':p.name,'confidence':p.confidence,'mapping':dict(p.mapping),'rows':p.rows_found,'evidence':p.evidence}
            for t_sheet in [s for s in spec['sheets'] if s.endswith('_ITEMS')]:
                if not item_candidates:
                    issues.append(_problem(t_sheet,0,'Sheet','No *_item sheet could be detected for this entity, so line-level quantities/rates are not validated yet; header rows are reviewed normally.','BLOCKED','Provide the item sheet in the workbook (one material per row) so line detail can be validated and later imported through the domain service.'))
                    continue
                for p in item_candidates:
                    p.target_sheet=t_sheet
                    for excel_row,data,notes in discovery.adapt_sheet_rows(kind, p, rows_provider, id_ref_maps, target_sheet=t_sheet):
                        errors=_profile_row(t_sheet,excel_row,data,kind,notes)
                        if p.confidence=='MEDIUM':
                            errors.append(_problem(t_sheet,excel_row,'Mapping',f"Sheet '{p.name}' was matched to {kind} items with MEDIUM confidence — row kept out of automatic import until reviewed.",'WARNING','Confirm the detected entity and mapping, then revalidate.'))
                        errors+=_check_references(kind,data,excel_row,t_sheet)
                        sheet_rows.append((t_sheet,excel_row,data,errors,p.name))
                    adapted_info[t_sheet]={'source_sheet':p.name,'confidence':p.confidence,'mapping':dict(p.mapping),'rows':p.rows_found,'evidence':p.evidence}
            issues.append(_problem(target_header,0,'Discovery','Workbook does not contain the official template sheets; legacy sheets were auto-detected and adapted. Review the mapping and every WARNING row before importing.','WARNING','Confirm mapping in the preview, then approve import.'))
        else:
            near_misses=[p for p in profiles if p.entity==kind and p.rows_found>0]
            for p in near_misses:
                issues.append(_problem(p.name,0,'Mapping',f'Sheet has {p.rows_found} data row(s) resembling {kind} at {p.confidence} confidence, but the column mapping is incomplete — nothing was imported from it.','BLOCKED','Open mapping review: at minimum an identifier/reference column and the entity name column must map to the template fields.'))
            if not near_misses:
                for p in profiles:
                    if p.status!='IGNORED' and p.rows_found:
                        issues.append(_problem(p.name,0,'Mapping',f'Not imported: {p.reason}','BLOCKED',f'Sheet has {p.rows_found} real data row(s) but was not mapped to {kind} (detected entity: {p.entity}, confidence: {p.confidence}).'))
                issues.append(_problem('—',0,'Template',f'No sheet in this workbook matches the {kind} template and no confidently detectable {kind} data was found. All {len([p for p in profiles if p.rows_found])} non-empty sheet(s) are listed in the sheet report with counts and reasons.','BLOCKED','Use AUTO analysis to see detected entities, download the official template, or extend the mapping configuration.'))
    stored=_save_source(raw, str(uuid.uuid4()))
    extra={'STORED_SOURCE':stored}
    if mode=='LEGACY_ADAPTED': extra['ADAPTATION']=adapted_info
    run=_persist_run(kind, raw, filename, actor, sheet_rows, issues, profiles, mode, extra)
    return run, issues

# ---------------------------------------------------------------------------
# AUTO analysis (wizard steps 1–3): workbook-level discovery report
# ---------------------------------------------------------------------------

def analyze_upload(raw, filename, actor=''):
    try: wb=load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc: raise ValueError('The file is not a readable XLSX workbook.') from exc
    profiles=discovery.profile_workbook(wb)
    discovery.detect_entities(profiles)
    order=discovery.recommended_order(profiles)
    entities={}
    for p in profiles:
        if p.status!='MAPPED' or p.entity not in TEMPLATES: continue
        e=entities.setdefault(p.entity,{'sheets':[],'rows':0,'confidences':[],'mapping':dict(p.mapping)})
        e['sheets'].append(p.name); e['rows']+=p.rows_found; e['confidences'].append(p.confidence)
    for e in entities.values():
        e['confidence']='HIGH' if 'HIGH' in e['confidences'] else 'MEDIUM'
    total_found=sum(p.rows_found for p in profiles)
    unmapped=sum(p.rows_found for p in profiles if p.status=='NOT_MAPPED')
    stored=_save_source(raw, str(uuid.uuid4()))
    run=MigrationRun(run_key=str(uuid.uuid4()),template_type='ANALYSIS',filename=filename,source_hash=filehash(raw),status='DISCOVERED',uploaded_by=actor,mode='DISCOVERY',summary_json='{}')
    db.session.add(run); db.session.flush()
    summary={'SOURCE ROWS FOUND':total_found,'MAPPED':sum(e['rows'] for e in entities.values()),'UNMAPPED':unmapped,
             'SHEETS':[p.as_dict() for p in profiles],'DETECTED_ENTITIES':entities,'IMPORT ORDER':order,'STORED_SOURCE':stored,
             'Total Rows':total_found,'READY':0,'WARNING':0,'INVALID':0,'EXACT_DUPLICATE':0,'ORPHAN':0,
             'BLOCKED':unmapped,'IMPORT ENABLED':False,
             'NEXT STEP':'Create a validation run per entity below, import masters in the listed order, then prepare transactions.'}
    run.summary_json=json.dumps(summary,default=str); db.session.commit()
    return run

def rerun_prepare(run_id, kind, actor=''):
    """Wizard step: build a real validation run for one detected entity from the
    preserved original workbook (no re-upload needed)."""
    run=MigrationRun.query.get_or_404(run_id)
    if kind not in TEMPLATES: raise ValueError('Unknown migration template.')
    raw=_load_source(run)
    if raw is None: raise ValueError('The preserved source workbook is no longer on disk; re-upload the file.')
    child, issues = validate_upload(kind, raw, run.filename, actor)
    summary=json.loads(child.summary_json or '{}')
    summary['Prepared from analysis run']=run_id
    child.summary_json=json.dumps(summary,default=str)
    db.session.commit()
    return child, issues

# ---------------------------------------------------------------------------
# Controlled import (masters only; transactions stay behind domain services)
# ---------------------------------------------------------------------------

def import_run(run):
    # Only low-risk master data is presently enabled. Transaction templates remain validation/dry-run only,
    # protecting stock/ledger state until a domain-service adapter is approved for each transaction type.
    if run.template_type not in MASTER_KINDS: raise ValueError('This transaction template is validation-ready but import is locked pending its domain-service adapter; no records were changed.')
    summary=json.loads(run.summary_json or '{}')
    model={'CLIENTS':Client,'SUPPLIERS':Supplier,'MATERIALS':Material,'ACCOUNTS':Account}[run.template_type]
    before_total=model.query.count()
    ready_before=sum(1 for r in run.rows if r.status=='READY')
    created=0
    for row in run.rows:
        if row.status!='READY': continue
        d=json.loads(row.data_json); ref=row.legacy_reference
        if MigrationMapping.query.filter_by(template_type=run.template_type,legacy_reference=ref).first():
            row.status='EXACT_DUPLICATE'
            summary['EXACT_DUPLICATE']=summary.get('EXACT_DUPLICATE',0)+1
            summary['READY']=max(0,summary.get('READY',1)-1)
            continue
        if run.template_type=='CLIENTS':
            obj=Client(code='MIG-'+str(uuid.uuid4())[:8].upper(),name=d['Client Name'],phone=d['Phone'] or None,address=d['Address'] or None,category=d['Category'] or 'General')
        elif run.template_type=='SUPPLIERS': obj=Supplier(name=d['Supplier Name'],phone=d['Phone'] or None,address=d['Address'] or None)
        elif run.template_type=='MATERIALS':
            cat=None
            cat_name=(d.get('Category') or '').strip()
            if cat_name:
                cat=MaterialCategory.query.filter(func.lower(MaterialCategory.name)==norm(cat_name)).first()
                if not cat:
                    cat=MaterialCategory(name=cat_name, is_active=True)
                    db.session.add(cat)
                    db.session.flush()
            obj=Material(code='MIG-'+str(uuid.uuid4())[:8].upper(),name=d['Material Name'],category_id=cat.id if cat else None,unit=d['Unit'] or 'Bags',unit_price=float(str(d['Unit Price'] or 0).replace(',','') or 0))
        else: obj=Account(name=d['Account Name'],type=d['Account Type'] or 'General',account_type=d['Account Type'] or 'General',category=d['Category'] or 'General',balance=float(str(d['Opening Balance'] or 0).replace(',','') or 0),opening_balance=float(str(d['Opening Balance'] or 0).replace(',','') or 0),bank_name=d['Bank Name'] or None,account_number=d['Account Number'] or None)
        db.session.add(obj); db.session.flush(); row.status='IMPORTED'; row.entity_type=run.template_type[:-1].title(); row.new_entity_id=obj.id; row.imported_at=datetime.utcnow(); db.session.add(MigrationMapping(template_type=run.template_type,legacy_reference=ref,entity_type=row.entity_type,entity_id=obj.id,run_id=run.id)); created+=1
    after_total=model.query.count()
    summary['Created']=created
    summary['IMPORTED']=summary.get('IMPORTED',0)+created
    imported_total=sum(1 for r in run.rows if r.status=='IMPORTED')
    summary['RECONCILE']={'READY at validation':ready_before,'Imported this run':created,'Skipped duplicate at import':ready_before-created,
                          'Target table':run.template_type[:-1].title(),'Target count before':before_total,'Target count after':after_total,
                          'Created delta':after_total-before_total,'Balanced':(after_total-before_total)==imported_total,
                          'Not imported (needs review)':sum(1 for r in run.rows if r.status in ('WARNING','INVALID','ORPHAN','BLOCKED','EXACT_DUPLICATE') )}
    run.status='COMPLETED'; run.imported_at=datetime.utcnow(); run.summary_json=json.dumps(summary,default=str); db.session.commit(); return created
